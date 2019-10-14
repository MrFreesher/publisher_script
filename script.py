import openpyxl
import json
import sys

if len(sys.argv) < 2:
    print("Execute with parameters")
else:
    inputFilePath = sys.argv[1]
    outputFilePath = sys.argv[2]
    startRow = 3
    workbook_obj = openpyxl.load_workbook(inputFilePath)
    sheet_obj = workbook_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    categories = []
    indexWithoutCategory = 0
    for i in range(1, max_col):
        cell_obj = sheet_obj.cell(row=3, column=i)
        if cell_obj.value is not None:
            categories.append(cell_obj.value)
        else:
            indexWithoutCategory += 1
    fields = []
    for i in range(2, indexWithoutCategory + 1):
        field = sheet_obj.cell(row=4, column=i).value
        if field in fields:
            fields.append("{}_2".format(field))
        else:
            fields.append(field)
    data = []
    for i in range(5, max_row):
        field_counter = 0
        row_obj = {}
        categoryIndex = 0
        for j in range(2, indexWithoutCategory + 1):
            row_obj[fields[field_counter]] = sheet_obj.cell(row=i, column=j).value
            field_counter += 1
        rowCategories = []
        for j in range(indexWithoutCategory + 1, max_col):
            temp_cell = sheet_obj.cell(row=i, column=j).value
            if temp_cell is "x":
                rowCategories.append(categories[categoryIndex])
            categoryIndex += 1
        row_obj["dziedziny"] = rowCategories
        data.append(row_obj)
    json_content = json.dumps(data, indent=4, sort_keys=True)
    file = open(outputFilePath, "w", encoding="utf-8")
    file.write(json_content)

